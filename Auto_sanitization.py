import os
import csv
import time
import pandas as pd
from openpyxl import load_workbook

# ===============================================
# Character Shifting Maps (Forward and Backward)
# ===============================================
_forward_map = {}
_backward_map = {}

# For lowercase letters a-z
for i in range(ord('a'), ord('z') + 1):
    c = chr(i)
    nxt_c = chr(i + 1) if i < ord('z') else 'a'
    _forward_map[c] = nxt_c
    _backward_map[nxt_c] = c

# For uppercase letters A-Z
for i in range(ord('A'), ord('Z') + 1):
    c = chr(i)
    nxt_c = chr(i + 1) if i < ord('Z') else 'A'
    _forward_map[c] = nxt_c
    _backward_map[nxt_c] = c

# For digits 0-9
digits = '0123456789'
for i, c in enumerate(digits):
    nxt_c = digits[(i + 1) % 10]  # 9 wraps to 0
    _forward_map[c] = nxt_c
    _backward_map[nxt_c] = c

def shift_char_forward(c: str) -> str:
    """
    Shift a character forward using the forward map.
    The underscore '_' remains unchanged.
    """
    if c == '_':
        return c
    return _forward_map.get(c, c)

def shift_char_backward(c: str) -> str:
    """
    Shift a character backward using the backward map.
    The underscore '_' remains unchanged.
    """
    if c == '_':
        return c
    return _backward_map.get(c, c)

# ============================================
# Common Edge Character Shifting Function
# ============================================
def shift_edges_chars(s: str, shift_func, count: int = 2) -> str:
    """
    Shift the first and last 'count' non-space characters in a string
    using the provided shift function.
    
    :param s: Input string
    :param shift_func: Function to apply (e.g., shift_char_forward or shift_char_backward)
    :param count: Number of non-space characters to shift from each end (default is 2)
    :return: Modified string
    """
    arr = list(s)
    length = len(arr)
    
    # Shift first 'count' non-space characters
    shifted = 0
    i = 0
    while shifted < count and i < length:
        if arr[i] != ' ':
            arr[i] = shift_func(arr[i])
            shifted += 1
        i += 1

    # Shift last 'count' non-space characters
    shifted = 0
    i = length - 1
    while shifted < count and i >= 0:
        if arr[i] != ' ':
            arr[i] = shift_func(arr[i])
            shifted += 1
        i -= 1

    return ''.join(arr)

# ============================================
# Sanitization and Desanitization Functions
# ============================================
def sanitize(s: str) -> str:
    """
    Sanitize a string by shifting the first and last two non-space characters forward.
    If the string is None, empty, "nan", or "none", return "N/A".
    
    :param s: Input string
    :return: Sanitized string
    """
    if s is None:
        return "N/A"
    s = str(s)
    if not s.strip() or s.lower() in ("nan", "none"):
        return "N/A"
    return shift_edges_chars(s, shift_char_forward)

def desanitize(s: str) -> str:
    """
    Desanitize a string by shifting the first and last two non-space characters backward.
    If the string is None, return an empty string.
    
    :param s: Input string
    :return: Desanitized string
    """
    if s is None:
        return ""
    s = str(s)
    return shift_edges_chars(s, shift_char_backward)

# ============================================
# Memoization Functions for Efficiency
# ============================================
def memo_sanitize(val, cache):
    """
    Cache the result of the sanitize function to avoid redundant computation.
    
    :param val: Input value
    :param cache: Dictionary for caching
    :return: Sanitized result
    """
    if val in cache:
        return cache[val]
    result = sanitize(val)
    cache[val] = result
    return result

def memo_desanitize(val, cache):
    """
    Cache the result of the desanitize function to avoid redundant computation.
    
    :param val: Input value
    :param cache: Dictionary for caching
    :return: Desanitized result
    """
    if val in cache:
        return cache[val]
    result = desanitize(val)
    cache[val] = result
    return result

# @EW, Kinly modify here for the new column to be sanitized
# ============================================
# Global Configuration for Sanitization Columns
# ============================================
# Modify this dictionary to adjust which columns should be sanitized
# and their corresponding output column names.
SANITIZATION_COLUMNS = {
    "external part": "external_part_sanitized",
    "ex_part": "external_part_sanitized",
    "internal part": "internal_part_sanitized",
    "internal part(old)": "internal part(old)_sanitized",
    "internal part(new)": "internal part(new)_sanitized"
}

def get_sanitized_column_name(original: str) -> str:
    """
    Get the sanitized column name based on the original column name.
    If the lowercased original exists in SANITIZATION_COLUMNS, return the mapped name.
    Otherwise, append "_sanitized" to the original.
    
    :param original: Original column name
    :return: New sanitized column name
    """
    return SANITIZATION_COLUMNS.get(original.lower(), f"{original}_sanitized")

# ============================================
# DataFrame Column Processing Functions
# ============================================
def insert_sanitized_columns(df: pd.DataFrame, columns_to_add: list):
    """
    Replace specified columns in the DataFrame with their sanitized versions.
    The new column names are determined by get_sanitized_column_name.
    
    :param df: Input DataFrame
    :param columns_to_add: List of column names to sanitize
    """
    # Process columns in reverse order to maintain indices when dropping/inserting.
    columns_with_index = [(col, df.columns.get_loc(col)) for col in columns_to_add]
    columns_with_index.sort(key=lambda x: x[1], reverse=True)
    
    for col, idx in columns_with_index:
        series = df[col].fillna('N/A') \
                      .replace(r'^\s*$', 'N/A', regex=True) \
                      .replace(['none', 'nan'], 'N/A')
        cache = {}
        processed_series = series.apply(lambda x: memo_sanitize(x, cache))
        df.drop(columns=[col], inplace=True)
        new_col_name = get_sanitized_column_name(col)
        df.insert(idx, new_col_name, processed_series)

def insert_desanitized_columns(df: pd.DataFrame, columns_to_process: list):
    """
    Replace specified columns (expected to be 'External part ID') in the DataFrame with their desanitized versions.
    The new column will be named "External Part".
    
    :param df: Input DataFrame
    :param columns_to_process: List of column names to desanitize
    """
    columns_with_index = [(col, df.columns.get_loc(col)) for col in columns_to_process]
    columns_with_index.sort(key=lambda x: x[1], reverse=True)
    
    for col, idx in columns_with_index:
        if col == 'External part ID':
            cache = {}
            processed_series = df[col].apply(lambda x: memo_desanitize(x, cache))
            df.drop(columns=[col], inplace=True)
            df.insert(idx, 'External Part', processed_series)

# ============================================
# XLSX Streaming Processing Functions
# ============================================
def process_xlsx_sanitization_streaming(input_path: str, output_folder: str):
    """
    Process an .xlsx file in streaming mode and sanitize specified columns.
    Columns to be sanitized are determined by SANITIZATION_COLUMNS (case-insensitive).
    The sanitized output replaces the original columns.
    
    Output file naming:
      - Single sheet: baseName_sanitized.csv
      - Multiple sheets: baseName_sheetName_sanitized.csv
    
    :param input_path: Path to the input .xlsx file
    :param output_folder: Folder to save the sanitized CSV output
    """
    base_name = os.path.splitext(os.path.basename(input_path))[0]
    workbook = load_workbook(filename=input_path, read_only=True, data_only=True)
    
    valid_columns = set(SANITIZATION_COLUMNS.keys())
    sheet_names = workbook.sheetnames
    total_sheets = len(sheet_names)
    
    for sheet_index, sheet_name in enumerate(sheet_names, start=1):
        worksheet = workbook[sheet_name]
        out_filename = f"{base_name}_sanitized.csv" if total_sheets == 1 else f"{base_name}_{sheet_name}_sanitized.csv"
        out_path = os.path.join(output_folder, out_filename)
        
        print(f"[SANITIZE] Processing sheet {sheet_index}/{total_sheets}: {sheet_name} -> {out_filename}")
        
        with open(out_path, mode='w', newline='', encoding='utf-8-sig') as f_out:
            writer = csv.writer(f_out)
            cache_dict = {}
            processed_indices = set()
            
            for row_index, row in enumerate(worksheet.iter_rows(values_only=True)):
                row_list = list(row) if row else []
                
                if row_index == 0:
                    new_header = []
                    for col_index, col_name in enumerate(row_list):
                        if col_name is not None and col_name.lower() in valid_columns:
                            new_header.append(get_sanitized_column_name(col_name))
                            processed_indices.add(col_index)
                        else:
                            new_header.append(col_name)
                    writer.writerow(new_header)
                else:
                    new_row = []
                    for col_index, value in enumerate(row_list):
                        if col_index in processed_indices:
                            new_row.append(memo_sanitize(value, cache_dict))
                        else:
                            new_row.append(value)
                    writer.writerow(new_row)
    workbook.close()

def process_xlsx_desanitization_streaming(input_path: str, output_folder: str):
    """
    Process an .xlsx file in streaming mode and desanitize the 'External part ID' column.
    The desanitized output replaces the original column, with the column name set to "External Part".
    
    Output file naming:
      - Single sheet: baseName_desanitized.csv
      - Multiple sheets: baseName_sheetName_desanitized.csv
    
    :param input_path: Path to the input .xlsx file
    :param output_folder: Folder to save the desanitized CSV output
    """
    base_name = os.path.splitext(os.path.basename(input_path))[0]
    workbook = load_workbook(filename=input_path, read_only=True, data_only=True)
    sheet_names = workbook.sheetnames
    total_sheets = len(sheet_names)
    
    for sheet_index, sheet_name in enumerate(sheet_names, start=1):
        worksheet = workbook[sheet_name]
        out_filename = f"{base_name}_desanitized.csv" if total_sheets == 1 else f"{base_name}_{sheet_name}_desanitized.csv"
        out_path = os.path.join(output_folder, out_filename)
        
        print(f"[DESANITIZE] Processing sheet {sheet_index}/{total_sheets}: {sheet_name} -> {out_filename}")
        
        with open(out_path, mode='w', newline='', encoding='utf-8-sig') as f_out:
            writer = csv.writer(f_out)
            cache_dict = {}
            processed_indices = set()
            
            for row_index, row in enumerate(worksheet.iter_rows(values_only=True)):
                row_list = list(row) if row else []
                if row_index == 0:
                    new_header = []
                    for col_index, col_name in enumerate(row_list):
                        if col_name == 'External part ID':
                            new_header.append("External Part")
                            processed_indices.add(col_index)
                        else:
                            new_header.append(col_name)
                    writer.writerow(new_header)
                else:
                    new_row = []
                    for col_index, value in enumerate(row_list):
                        if col_index in processed_indices:
                            new_row.append(memo_desanitize(value, cache_dict))
                        else:
                            new_row.append(value)
                    writer.writerow(new_row)
    workbook.close()

# ============================================
# File Processing Functions
# ============================================
def process_file_sanitization(file_path: str, output_folder: str):
    """
    Process a file (either .xlsx or .csv) for sanitization.
    The specified columns are replaced with their sanitized versions.
    
    :param file_path: Path to the input file
    :param output_folder: Folder to save the sanitized output
    """
    base_name, ext = os.path.splitext(os.path.basename(file_path))
    ext = ext.lower()
    
    if ext == '.xlsx':
        process_xlsx_sanitization_streaming(file_path, output_folder)
    elif ext == '.csv':
        try:
            df = pd.read_csv(file_path, encoding='utf-8')
        except Exception:
            df = pd.read_csv(file_path, encoding='gbk')
        
        columns_to_add = [col for col in df.columns if col.lower() in SANITIZATION_COLUMNS]
        if columns_to_add:
            insert_sanitized_columns(df, columns_to_add)
        
        out_filename = f"{base_name}_sanitized.csv"
        out_path = os.path.join(output_folder, out_filename)
        df.to_csv(out_path, index=False, encoding='utf-8-sig')

def process_file_desanitization(file_path: str, output_folder: str):
    """
    Process a file (either .xlsx or .csv) for desanitization.
    The 'External part ID' column is replaced with its desanitized version.
    
    :param file_path: Path to the input file
    :param output_folder: Folder to save the desanitized output
    """
    base_name, ext = os.path.splitext(os.path.basename(file_path))
    ext = ext.lower()
    
    if ext == '.xlsx':
        process_xlsx_desanitization_streaming(file_path, output_folder)
    elif ext == '.csv':
        try:
            df = pd.read_csv(file_path, encoding='utf-8')
        except Exception:
            df = pd.read_csv(file_path, encoding='gbk')
        
        columns_to_process = []
        if 'External part ID' in df.columns:
            columns_to_process.append('External part ID')
        
        if columns_to_process:
            insert_desanitized_columns(df, columns_to_process)
        else:
            print(f"The file {file_path} does not contain sanitized columns.")
        
        out_filename = f"{base_name}_desanitized.csv"
        out_path = os.path.join(output_folder, out_filename)
        df.to_csv(out_path, index=False, encoding='utf-8-sig')

# ============================================
# Main Workflow Functions
# ============================================
def sanitize_data():
    """
    Process all .xlsx and .csv files in the 'Unsanitized' folder,
    perform sanitization, and output the results to the 'Sanitized' folder.
    """
    print("\nSanitizing process started...")
    
    input_folder = 'Unsanitized'
    output_folder = 'Sanitized'
    os.makedirs(output_folder, exist_ok=True)
    
    file_list = [f for f in os.listdir(input_folder) if f.endswith(('.xlsx', '.csv'))]
    if not file_list:
        print("No .xlsx or .csv files found in 'Unsanitized' folder.")
        return
    
    print("\nFiles to be sanitized:")
    for file_name in file_list:
        print(f" - {file_name}")
    
    for file_name in file_list:
        file_path = os.path.join(input_folder, file_name)
        print(f"Processing file: {file_name}")
        process_file_sanitization(file_path, output_folder)
    
    print("\nSanitizing process completed.")

def desanitize_data():
    """
    Process all .xlsx and .csv files in the 'Undesanitized' folder,
    perform desanitization, and output the results to the 'Desanitized' folder.
    """
    print("\nDesanitizing process started...")
    
    input_folder = 'Undesanitized'
    output_folder = 'Desanitized'
    os.makedirs(output_folder, exist_ok=True)
    
    file_list = [f for f in os.listdir(input_folder) if f.endswith(('.xlsx', '.csv'))]
    if not file_list:
        print("No .xlsx or .csv files found in 'Undesanitized' folder.")
        return
    
    print("\nFiles to be desanitized:")
    for file_name in file_list:
        print(f" - {file_name}")
    
    for file_name in file_list:
        file_path = os.path.join(input_folder, file_name)
        process_file_desanitization(file_path, output_folder)
    
    print("\nDesanitizing process completed.")

def main():
    """
    Main function: Ask the user if they want to sanitize data.
    Depending on the input, perform sanitization or desanitization.
    Display the total processing time.
    """
    print("Want to sanitize data? (y/n)")
    user_input = input().strip().lower()
    
    start_time = time.time()
    
    if user_input == "y":
        sanitize_data()
    elif user_input == "n":
        desanitize_data()
    else:
        print("Invalid input, please enter 'y' or 'n'.")
    
    elapsed_minutes = round((time.time() - start_time) / 60, 2)
    print(f"\nTime taken: {elapsed_minutes} minutes.")
    input("\nProcessing complete. Press Enter to exit...")

if __name__ == "__main__":
    main()